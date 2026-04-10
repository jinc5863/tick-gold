import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
from jinja2 import Template
import os

class ComplianceReporter:
    """合规报告类"""

    def __init__(self, report_directory: str = "reports"):
        self.report_directory = report_directory
        self.os.makedirs(report_directory, exist_ok=True)

    def generate_daily_report(self, trading_data: dict, risk_metrics: dict, audit_results: dict) -> str:
        """生成每日合规报告"""
        report_data = {
            'report_date': datetime.now().strftime('%Y-%m-%d'),
            'generation_time': datetime.now().isoformat(),
            'trading_summary': self._generate_trading_summary(trading_data),
            'risk_metrics': risk_metrics,
            'audit_results': audit_results,
            'compliance_status': self._check_compliance_status(risk_metrics, audit_results)
        }

        # 生成PDF报告
        pdf_path = os.path.join(self.report_directory, f"compliance_report_{report_data['report_date']}.pdf")
        self._generate_pdf_report(report_data, pdf_path)

        # 生成Excel报告
        excel_path = os.path.join(self.report_directory, f"compliance_report_{report_data['report_date']}.xlsx")
        self._generate_excel_report(report_data, excel_path)

        return pdf_path, excel_path

    def _generate_trading_summary(self, trading_data: dict) -> dict:
        """生成交易摘要"""
        if 'trades' not in trading_data or 'equity' not in trading_data:
            return {}

        trades = trading_data['trades']
        equity = trading_data['equity']

        return {
            'total_trades': len(trades),
            'winning_trades': sum(1 for t in trades if t['profit'] > 0),
            'losing_trades': sum(1 for t in trades if t['profit'] < 0),
            'total_profit': sum(t['profit'] for t in trades),
            'average_profit': np.mean([t['profit'] for t in trades]) if trades else 0,
            'final_equity': equity.iloc[-1] if not equity.empty else 0,
            'daily_return': (equity.iloc[-1] - equity.iloc[0]) / equity.iloc[0] if len(equity) > 1 else 0
        }

    def _check_compliance_status(self, risk_metrics: dict, audit_results: dict) -> dict:
        """检查合规状态"""
        status = {
            'overall_status': 'passed',
            'risk_status': 'passed',
            'audit_status': 'passed',
            'issues': []
        }

        # 检查风险指标
        if risk_metrics['drawdown'] > 0.02:  # 超过2%回撤
            status['risk_status'] = 'failed'
            status['overall_status'] = 'failed'
            status['issues'].append(f"回撤超过限制: {risk_metrics['drawdown']:.2%}")

        if risk_metrics['daily_loss'] < -0.005:  # 超过0.5%当日亏损
            status['risk_status'] = 'failed'
            status['overall_status'] = 'failed'
            status['issues'].append(f"当日亏损超过限制: {risk_metrics['daily_loss']:.2%}")

        # 检查审计结果
        if audit_results.get('audit_items', {}).get('duplicates', 0) > 0:
            status['audit_status'] = 'warning'
            status['issues'].append(f"发现重复数据: {audit_results['audit_items']['duplicates']}条")

        if audit_results.get('audit_items', {}).get('abnormal_spreads_count', 0) > 0:
            status['audit_status'] = 'warning'
            status['issues'].append(f"发现异常点差: {audit_results['audit_items']['abnormal_spreads_count']}条")

        return status

    def _generate_pdf_report(self, report_data: dict, output_path: str):
        """生成PDF报告"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors

            doc = SimpleDocTemplate(output_path, pagesize=letter)
            styles = getSampleStyleSheet()
            elements = []

            # 标题
            title = Paragraph(f"合规报告 - {report_data['report_date']}", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))

            # 交易摘要
            elements.append(Paragraph("交易摘要", styles['Heading2']))
            trade_summary = report_data['trading_summary']
            trade_table = [
                ['总交易次数', str(trade_summary['total_trades'])],
                ['盈利交易', str(trade_summary['winning_trades'])],
                ['亏损交易', str(trade_summary['losing_trades'])],
                ['总盈利', f"${trade_summary['total_profit']:,.2f}"],
                ['平均盈利', f"${trade_summary['average_profit']:,.2f}"],
                ['最终权益', f"${trade_summary['final_equity']:,.2f}"],
                ['日收益率', f"{trade_summary['daily_return']:.2%}"]
            ]
            table = Table(trade_table)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            elements.append(Spacer(1, 12))

            # 风险指标
            elements.append(Paragraph("风险指标", styles['Heading2']))
            risk_table = [
                ['回撤', f"{report_data['risk_metrics']['drawdown']:.2%}"],
                ['当日亏损', f"{report_data['risk_metrics']['daily_loss']:.2%}"],
                ['仓位风险', f"{report_data['risk_metrics']['position_risk']:.2%}"],
                ['跳空风险', f"{report_data['risk_metrics']['gap_risk']:.2%}"],
                ['隔夜风险', f"{report_data['risk_metrics']['overnight_risk']:.2%}"]
            ]
            table = Table(risk_table)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            elements.append(Spacer(1, 12))

            # 合规状态
            elements.append(Paragraph("合规状态", styles['Heading2']))
            status = report_data['compliance_status']
            status_table = [
                ['总体状态', status['overall_status']],
                ['风险状态', status['risk_status']],
                ['审计状态', status['audit_status']]
            ]
            table = Table(status_table)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)

            # 问题列表
            if status['issues']:
                elements.append(Paragraph("发现的问题", styles['Heading3']))
                for issue in status['issues']:
                    elements.append(Paragraph(f"• {issue}", styles['BodyText']))

            doc.build(elements)
        except ImportError:
            print("ReportLab not installed, skipping PDF generation")

    def _generate_excel_report(self, report_data: dict, output_path: str):
        """生成Excel报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "合规报告"

            # 标题
            ws.merge_cells('A1:G1')
            ws['A1'] = f"合规报告 - {report_data['report_date']}"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            # 交易摘要
            ws.append(["交易摘要", "", "", "", "", "", ""])
            ws.append(["总交易次数", "盈利交易", "亏损交易", "总盈利", "平均盈利", "最终权益", "日收益率"])
            trade_summary = report_data['trading_summary']
            ws.append([
                trade_summary['total_trades'],
                trade_summary['winning_trades'],
                trade_summary['losing_trades'],
                trade_summary['total_profit'],
                trade_summary['average_profit'],
                trade_summary['final_equity'],
                trade_summary['daily_return']
            ])

            # 风险指标
            ws.append(["", "", "", "", "", "", ""])
            ws.append(["风险指标", "", "", "", "", "", ""])
            ws.append(["回撤", "当日亏损", "仓位风险", "跳空风险", "隔夜风险", "", ""])
            risk_metrics = report_data['risk_metrics']
            ws.append([
                risk_metrics['drawdown'],
                risk_metrics['daily_loss'],
                risk_metrics['position_risk'],
                risk_metrics['gap_risk'],
                risk_metrics['overnight_risk'],
                "",
                ""
            ])

            # 合规状态
            ws.append(["", "", "", "", "", "", ""])
            ws.append(["合规状态", "", "", "", "", "", ""])
            ws.append(["总体状态", "风险状态", "审计状态", "", "", "", ""])
            status = report_data['compliance_status']
            ws.append([
                status['overall_status'],
                status['risk_status'],
                status['audit_status'],
                "",
                "",
                "",
                ""
            ])

            # 问题列表
            if status['issues']:
                ws.append(["", "", "", "", "", "", ""])
                ws.append(["发现的问题", "", "", "", "", "", ""])
                for issue in status['issues']:
                    ws.append([issue, "", "", "", "", "", ""])

            wb.save(output_path)
        except ImportError:
            print("openpyxl not installed, skipping Excel generation")

    def generate_regulatory_report(self, report_data: dict) -> str:
        """生成监管报告"""
        # 这里应该生成符合监管要求的报告格式
        # 简化版本：生成JSON格式的监管报告
        regulatory_report = {
            'report_id': f"REG-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            'report_date': report_data['report_date'],
            'trading_summary': report_data['trading_summary'],
            'risk_metrics': report_data['risk_metrics'],
            'compliance_status': report_data['compliance_status'],
            'audit_results': report_data['audit_results']
        }

        json_path = os.path.join(self.report_directory, f"regulatory_report_{report_data['report_date']}.json")
        with open(json_path, 'w') as f:
            json.dump(regulatory_report, f, indent=2)

        return json_path